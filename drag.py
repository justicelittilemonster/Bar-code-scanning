coding ="UTF-8"

from urllib import    request, parse
import json
import sys
from openpyxl import load_workbook
import time
import winsound
import ctypes

count_col=1  #
num_ini=1 #初始药品数量
try:
    excel_name=sys.argv[1]  #获取需要存放的表格文件名
except :
    ctypes.windll.user32.MessageBoxW(0, "请将文件拖到本软件上打开", "错误", 0)
same=False #标记位  用来检测是否为相同药品

##################初始化表格######################
try:
    wb=load_workbook(excel_name)
except IOError:
    ctypes.windll.user32.MessageBoxW(0,"打开文件错误，请确保文件未被占用","错误",0)
    exit()
wb.template=False
ws= wb.active
count_row=ws.max_row+1; #获取最下面的一行
ws.cell(1,1).value="条形码编码"
ws.cell(1,2).value="药品名称"
ws.cell(1,3).value="生产厂家"
ws.cell(1,4).value="数量"
ws.cell(1,5).value="入库时间"
ws.cell(1,6).value="生产批号"
ws.cell(1,7).value="失效日期"

try:
    wb.save(excel_name)
except IOError:
    ctypes.windll.user32.MessageBoxW(0,"打开文件错误，请确保文件未被占用","错误",0)
    exit()


while True:
    print('请扫描条形码')
    code=input()
    showapi_appid="56508" 
    showapi_sign="ebf24ccfe3d049efb68f8a7fcd9f38e3"  
    url="http://route.showapi.com/66-22"
    send_data = parse.urlencode([
    ('showapi_appid', showapi_appid)
    ,('showapi_sign', showapi_sign)
	,('code', code)

    ])

    req = request.Request(url)
    try:
        response = request.urlopen(req, data=send_data.encode('utf-8'), timeout = 10) # 10秒超时反馈
    except Exception as e:
        print(e)
    result = response.read().decode('utf-8')
    data = json.loads(result) #转换成字典


    count_col=1;
    goodsName=data.get("showapi_res_body").get("goodsName")
    manuName=data.get("showapi_res_body").get("manuName")
    day_time=time.strftime('%Y-%m-%d',time.localtime(time.time()))

    try:
        wb=load_workbook(excel_name)
    except IOError:
        ctypes.windll.user32.MessageBoxW(0,"打开文件错误，请确保文件未被占用","错误",0)
        exit()
    wb.template=False
    ws= wb.active

    for i in range(1,ws.max_row+1): #判断是否为相同药品
        if(code==ws.cell(i,1).value and day_time== ws.cell(i,5).value): 
            ws.cell(i,4).value+=1
            same=True
            winsound.Beep(600,500)#发出声音
            print(goodsName)
            
            try:
                wb.save(excel_name)
            except IOError:
                ctypes.windll.user32.MessageBoxW(0,"打开文件错误，请确保文件未被占用","错误",0)
                exit()
    if(same==True):
        same=False
        continue
    ws.cell(count_row,count_col).value=code
    count_col+=1
    ws.cell(count_row,count_col).value=goodsName
    count_col+=1
    ws.cell(count_row,count_col).value=manuName
    count_col+=1
    ws.cell(count_row,count_col).value=num_ini
    count_col+=1
    ws.cell(count_row,count_col).value=day_time
    count_row+=1
    try:
        wb.save(excel_name)
    except IOError:
        ctypes.windll.user32.MessageBoxW(0,"打开文件错误，请确保文件未被占用","错误",0)
        exit()
    winsound.Beep(600,500)#发出声音
    print(goodsName)
